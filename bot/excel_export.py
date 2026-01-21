"""Экспорт данных проекта в Excel."""

import logging
from io import BytesIO
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

logger = logging.getLogger(__name__)


def export_project_to_excel(
    project_name: str,
    project_address: str,
    budget: float,
    transactions: list[dict],
    progress_photos: list[dict] | None = None,
) -> BytesIO | None:
    """
    Экспортировать данные проекта в Excel файл.
    
    Args:
        project_name: Название проекта
        project_address: Адрес проекта
        budget: Бюджет проекта
        transactions: Список транзакций (дикты с amount, category, description, created_at)
        progress_photos: Опционально список фото отчетов
    
    Returns:
        BytesIO объект с Excel файлом или None при ошибке
    """
    if not HAS_EXCEL:
        logger.error("pandas и openpyxl не установлены. Установите: pip install pandas openpyxl")
        return None
    
    try:
        logger.info(f"📊 Создание Excel отчета для проекта: {project_name}")
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Информация о проекте
        ws["A1"] = "ОТЧЕТ ПО ПРОЕКТУ"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_alignment
        ws.merge_cells("A1:E1")
        
        ws["A2"] = f"Проект: {project_name}"
        ws["A3"] = f"Адрес: {project_address}"
        ws["A4"] = f"Бюджет: {budget:,.2f} BYN"
        ws["A5"] = f"Дата отчета: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Таблица транзакций
        row = 7
        headers = ["Дата", "Категория", "Описание", "Сумма (BYN)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        
        row += 1
        total_spent = 0
        
        for trans in transactions:
            ws.cell(row=row, column=1, value=trans.get("created_at", ""))
            ws.cell(row=row, column=2, value=trans.get("category", ""))
            ws.cell(row=row, column=3, value=trans.get("description", ""))
            ws.cell(row=row, column=4, value=trans.get("amount", 0))
            
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
            
            total_spent += float(trans.get("amount", 0))
            row += 1
        
        # Итоги
        ws.cell(row=row, column=3, value="ИТОГО:").font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_spent).font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        
        # Статистика
        row += 2
        ws.cell(row=row, column=1, value="СТАТИСТИКА").font = Font(bold=True, size=11)
        
        row += 1
        ws.cell(row=row, column=1, value="Потрачено:")
        ws.cell(row=row, column=2, value=f"{total_spent:,.2f} BYN")
        
        row += 1
        remaining = budget - total_spent
        ws.cell(row=row, column=1, value="Осталось:")
        ws.cell(row=row, column=2, value=f"{remaining:,.2f} BYN")
        
        row += 1
        percentage = (total_spent / budget * 100) if budget > 0 else 0
        ws.cell(row=row, column=1, value="% использовано:")
        ws.cell(row=row, column=2, value=f"{percentage:.1f}%")
        
        # Автоширина колонок
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"✅ Excel отчет создан успешно")
        return output
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel: {e}")
        return None


def export_project_summary(
    project_name: str,
    budget: float,
    total_spent: float,
    by_category: dict[str, float],
    transactions_count: int,
) -> BytesIO | None:
    """
    Экспортировать краткий отчет проекта.
    
    Args:
        project_name: Название проекта
        budget: Бюджет
        total_spent: Всего потрачено
        by_category: Расходы по категориям
        transactions_count: Количество транзакций
    
    Returns:
        BytesIO объект с Excel файлом
    """
    if not HAS_EXCEL:
        logger.error("pandas и openpyxl не установлены")
        return None
    
    try:
        logger.info(f"📊 Создание краткого отчета: {project_name}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводка"
        
        # Заголовок
        ws["A1"] = f"Сводка по проекту: {project_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")
        
        # Данные
        row = 3
        data = [
            ("Бюджет:", f"{budget:,.2f} BYN"),
            ("Потрачено:", f"{total_spent:,.2f} BYN"),
            ("Осталось:", f"{budget - total_spent:,.2f} BYN"),
            ("% использовано:", f"{(total_spent/budget*100) if budget > 0 else 0:.1f}%"),
            ("Количество расходов:", str(transactions_count)),
        ]
        
        for label, value in data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # По категориям
        row += 1
        ws.cell(row=row, column=1, value="По категориям:").font = Font(bold=True)
        row += 1
        
        for category, amount in by_category.items():
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=f"{amount:,.2f} BYN")
            row += 1
        
        # Ширина
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"✅ Краткий отчет создан")
        return output
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании отчета: {e}")
        return None
